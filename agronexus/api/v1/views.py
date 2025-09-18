"""
AgroNexus - Sistema 
ViewSets para API REST
"""

from datetime import timedelta

from django.core.exceptions import ValidationError
from django.db.models import Avg, Count, F, Q, Sum
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import OpenApiParameter, extend_schema
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.response import Response

from ...models import (AdministracaoMedicamento, Animal, AnimalManejo, Area,
                       CalendarioSanitario, CategoriaFinanceira,
                       ConfiguracaoSistema, ContaFinanceira,
                       DiagnosticoGestacao, EspecieAnimal, EstacaoMonta,
                       HistoricoLoteAnimal, HistoricoOcupacaoArea, Inseminacao,
                       LancamentoFinanceiro, Lote, Manejo, Medicamento, Parto,
                       Pesagem, Propriedade, ProtocoloIATF, RacaAnimal,
                       RelatorioPersonalizado, Usuario, Vacina, Vacinacao)
from ...permissions.base import IsOwnerOrReadOnly, PropriedadeOwnerPermission
from ...utils.filters import (AnimalFilter, AreaFilter,
                              CalendarioSanitarioFilter,
                              DiagnosticoGestacaoFilter, EstacaoMontaFilter,
                              InseminacaoFilter, LancamentoFinanceiroFilter,
                              LoteFilter, ManejoFilter, PartoFilter,
                              PesagemFilter, PropriedadeFilter,
                              VacinacaoFilter)
from .serializers import (AdministracaoMedicamentoSerializer, AnimalSerializer,
                          AreaSerializer, CalendarioSanitarioSerializer,
                          CategoriaFinanceiraSerializer,
                          ConfiguracaoSistemaSerializer,
                          ContaFinanceiraSerializer,
                          DiagnosticoGestacaoSerializer,
                          EspecieAnimalSerializer, EstacaoMontaSerializer,
                          HistoricoLoteAnimalSerializer,
                          HistoricoOcupacaoAreaSerializer,
                          InseminacaoSerializer,
                          LancamentoFinanceiroSerializer, LoteSerializer,
                          ManejoSerializer, MedicamentoSerializer,
                          PartoSerializer, PesagemSerializer,
                          PropriedadeSerializer, ProtocoloIATFSerializer,
                          RacaAnimalSerializer,
                          RelatorioPersonalizadoSerializer, UsuarioSerializer,
                          VacinacaoSerializer, VacinaSerializer)


class BaseViewSet(viewsets.ModelViewSet):
    """ViewSet base com funcionalidades comuns"""
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]

    def get_queryset(self):
        """Filtra registros por propriedade quando aplicável"""
        queryset = super().get_queryset()

        # Se o modelo tem propriedade, filtra por propriedades do usuário
        if hasattr(self.queryset.model, 'propriedade'):
            propriedades_usuario = Propriedade.objects.filter(
                proprietario=self.request.user
            )
            queryset = queryset.filter(propriedade__in=propriedades_usuario)

        return queryset

    def perform_create(self, serializer):
        """Associa o usuário atual ao criar registros"""
        # Se o modelo tem campo usuario, associa o usuário atual
        if hasattr(serializer.Meta.model, 'usuario'):
            serializer.save(usuario=self.request.user)
        else:
            serializer.save()


# ============================================================================
# VIEWSETS DE USUÁRIOS E PROPRIEDADES
# ============================================================================

class UsuarioViewSet(viewsets.ModelViewSet):
    """ViewSet para usuários"""
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    search_fields = ['username', 'first_name', 'last_name', 'email']
    ordering_fields = ['username', 'first_name', 'last_name', 'date_joined']
    ordering = ['-date_joined']

    def get_queryset(self):
        """Usuários podem ver apenas dados da sua propriedade"""
        if self.request.user.is_superuser:
            return self.queryset

        # Pega as propriedades do usuário atual
        propriedades = Propriedade.objects.filter(
            proprietario=self.request.user)

        # Retorna usuários das suas propriedades
        return self.queryset.filter(
            Q(propriedades__in=propriedades) | Q(id=self.request.user.id)
        ).distinct()

    @action(detail=False, methods=['get'])
    def me(self, request):
        """Retorna dados do usuário atual, incluindo propriedades"""
        serializer = self.get_serializer(request.user)
        # Busca propriedades do usuário
        propriedades = Propriedade.objects.filter(proprietario=request.user)
        propriedades_data = PropriedadeSerializer(propriedades, many=True).data
        data = serializer.data
        data["propriedades"] = propriedades_data
        return Response(data)


class PropriedadeViewSet(BaseViewSet):
    """ViewSet para propriedades"""
    queryset = Propriedade.objects.all()
    serializer_class = PropriedadeSerializer
    permission_classes = [permissions.IsAuthenticated, IsOwnerOrReadOnly]
    filterset_class = PropriedadeFilter
    search_fields = ['nome', 'localizacao', 'inscricao_estadual']
    ordering_fields = ['nome', 'area_total_ha', 'data_criacao']
    ordering = ['nome']

    def get_queryset(self):
        """Usuários veem apenas suas propriedades"""
        return self.queryset.filter(proprietario=self.request.user).annotate(
            area_ocupada=Sum('areas__tamanho_ha'),
            total_animais=Count('animais', filter=Q(animais__status='ativo')),
            total_lotes=Count('lotes', filter=Q(lotes__ativo=True)),
            total_areas=Count('areas')
        )

    def perform_create(self, serializer):
        """Associa o usuário atual como proprietário"""
        serializer.save(proprietario=self.request.user)

    @action(detail=True, methods=['get'])
    def dashboard(self, request, pk=None):
        """Dashboard com estatísticas da propriedade"""
        propriedade = self.get_object()

        # Estatísticas gerais
        total_animais = propriedade.animais.filter(status='ativo').count()
        total_lotes = propriedade.lotes.filter(ativo=True).count()
        total_areas = propriedade.areas.count()

        # Estatísticas por categoria
        por_categoria = propriedade.animais.filter(status='ativo').values(
            'categoria'
        ).annotate(total=Count('id'))

        # Estatísticas por sexo
        por_sexo = propriedade.animais.filter(status='ativo').values(
            'sexo'
        ).annotate(total=Count('id'))

        # Estatísticas financeiras (último mês)
        data_limite = timezone.now().date() - timedelta(days=30)
        receitas = propriedade.lancamentos_financeiros.filter(
            tipo='entrada',
            data_lancamento__gte=data_limite
        ).aggregate(total=Sum('valor'))['total'] or 0

        despesas = propriedade.lancamentos_financeiros.filter(
            tipo='saida',
            data_lancamento__gte=data_limite
        ).aggregate(total=Sum('valor'))['total'] or 0

        return Response({
            'estatisticas_gerais': {
                'total_animais': total_animais,
                'total_lotes': total_lotes,
                'total_areas': total_areas,
                'area_total_ha': propriedade.area_total_ha,
                'taxa_ocupacao_global': propriedade.get_taxa_ocupacao_global()
            },
            'distribuicao_animais': {
                'por_categoria': list(por_categoria),
                'por_sexo': list(por_sexo)
            },
            'financeiro_mensal': {
                'receitas': receitas,
                'despesas': despesas,
                'saldo': receitas - despesas
            }
        })


# ============================================================================
# VIEWSETS DE ÁREAS
# ============================================================================

class AreaViewSet(BaseViewSet):
    """ViewSet para áreas"""
    queryset = Area.objects.all()
    serializer_class = AreaSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]
    filterset_class = AreaFilter
    search_fields = ['nome', 'tipo_forragem']
    ordering_fields = ['nome', 'tipo', 'tamanho_ha', 'status']
    ordering = ['nome']

    def get_queryset(self):
        return super().get_queryset().select_related('propriedade')

    @action(detail=True, methods=['post'])
    def ocupar(self, request, pk=None):
        """Ocupa uma área com um lote"""
        area = self.get_object()
        lote_id = request.data.get('lote_id')

        if not lote_id:
            return Response(
                {'error': 'ID do lote é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            lote = Lote.objects.get(id=lote_id, propriedade=area.propriedade)
        except Lote.DoesNotExist:
            return Response(
                {'error': 'Lote não encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Verifica se a área já está ocupada
        if area.get_lote_atual():
            return Response(
                {'error': 'Área já está ocupada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Libera a área atual do lote
        if lote.area_atual:
            lote.area_atual.status = 'disponivel'
            lote.area_atual.save()

        # Ocupa a nova área
        lote.area_atual = area
        lote.save()

        area.status = 'em_uso'
        area.save()

        # Registra no histórico
        HistoricoOcupacaoArea.objects.create(
            lote=lote,
            area=area,
            data_entrada=timezone.now().date(),
            motivo_movimentacao=request.data.get(
                'motivo', 'Movimentação via API'),
            usuario=request.user
        )

        return Response({'message': 'Área ocupada com sucesso'})

    @action(detail=True, methods=['post'])
    def liberar(self, request, pk=None):
        """Libera uma área"""
        area = self.get_object()
        lote_atual = area.get_lote_atual()

        if not lote_atual:
            return Response(
                {'error': 'Área não está ocupada'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Finaliza o histórico atual
        historico = HistoricoOcupacaoArea.objects.filter(
            area=area,
            data_saida__isnull=True
        ).first()

        if historico:
            historico.data_saida = timezone.now().date()
            historico.motivo_movimentacao = request.data.get(
                'motivo', 'Liberação via API')
            historico.save()

        # Libera a área
        lote_atual.area_atual = None
        lote_atual.save()

        area.status = 'disponivel'
        area.save()

        return Response({'message': 'Área liberada com sucesso'})


# ============================================================================
# VIEWSETS DE ESPÉCIES E RAÇAS
# ============================================================================

class EspecieAnimalViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para espécies de animais (apenas leitura)"""
    queryset = EspecieAnimal.objects.filter(ativo=True)
    serializer_class = EspecieAnimalSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ['nome', 'nome_display']
    ordering_fields = ['nome', 'nome_display']
    ordering = ['nome']

    @action(detail=True, methods=['get'])
    def racas(self, request, pk=None):
        """Retorna todas as raças de uma espécie"""
        especie = self.get_object()
        racas = especie.racas.filter(ativo=True)
        serializer = RacaAnimalSerializer(racas, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def categorias(self, request, pk=None):
        """Retorna as categorias disponíveis para uma espécie"""
        especie = self.get_object()
        categorias = especie.get_categorias()
        return Response(categorias)


class RacaAnimalViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para raças de animais (apenas leitura)"""
    queryset = RacaAnimal.objects.filter(ativo=True)
    serializer_class = RacaAnimalSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['especie']
    search_fields = ['nome', 'origem']
    ordering_fields = ['nome', 'especie__nome']
    ordering = ['especie__nome', 'nome']


# ============================================================================
# VIEWSETS DE ANIMAIS E LOTES
# ============================================================================

class AnimalViewSet(BaseViewSet):
    """ViewSet para animais"""
    queryset = Animal.objects.all()
    serializer_class = AnimalSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]
    filterset_class = AnimalFilter
    search_fields = ['identificacao_unica', 'nome_registro', 'raca']
    ordering_fields = ['identificacao_unica',
                       'data_nascimento', 'categoria', 'data_criacao']
    ordering = ['identificacao_unica']

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'propriedade', 'lote_atual', 'pai', 'mae'
        )

        # Adiciona campos calculados
        for animal in queryset:
            animal.idade_dias = animal.get_idade_dias()
            animal.idade_meses = animal.get_idade_meses()
            animal.peso_atual = animal.get_peso_atual()
            animal.ua_value = animal.get_ua_value()

        return queryset

    @action(detail=True, methods=['get'])
    def historico_completo(self, request, pk=None):
        """Retorna histórico completo do animal"""
        animal = self.get_object()
        historico = animal.get_historico_completo()
        return Response(historico)

    @action(detail=True, methods=['get'])
    def evolucao_peso(self, request, pk=None):
        """Retorna evolução do peso do animal"""
        animal = self.get_object()
        pesagens = animal.pesagens.order_by('data_pesagem').values(
            'data_pesagem', 'peso_kg'
        )
        return Response(list(pesagens))

    @action(detail=True, methods=['post'])
    def trocar_lote(self, request, pk=None):
        """Troca o animal de lote"""
        animal = self.get_object()
        lote_id = request.data.get('lote_id')

        if not lote_id:
            return Response(
                {'error': 'ID do lote é obrigatório'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            novo_lote = Lote.objects.get(
                id=lote_id, propriedade=animal.propriedade)
        except Lote.DoesNotExist:
            return Response(
                {'error': 'Lote não encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        # Registra saída do lote atual
        if animal.lote_atual:
            historico_atual = HistoricoLoteAnimal.objects.filter(
                animal=animal,
                lote=animal.lote_atual,
                data_saida__isnull=True
            ).first()

            if historico_atual:
                historico_atual.data_saida = timezone.now().date()
                historico_atual.motivo_movimentacao = request.data.get(
                    'motivo', 'Troca via API')
                historico_atual.save()

        # Registra entrada no novo lote
        HistoricoLoteAnimal.objects.create(
            animal=animal,
            lote=novo_lote,
            data_entrada=timezone.now().date(),
            motivo_movimentacao=request.data.get('motivo', 'Troca via API'),
            usuario=request.user
        )

        animal.lote_atual = novo_lote
        animal.save()

        return Response({'message': 'Animal movido com sucesso'})

    @action(detail=False, methods=['post'])
    def exportar_excel(self, request):
        """
        Exporta animais para Excel com validações de segurança.
        Só permite exportar animais de propriedades que pertencem ao usuário autenticado.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        from django.utils import timezone
        import io

        # Validações de segurança
        propriedade_id = request.data.get('propriedade_id')
        incluir_genealogia = request.data.get('incluir_genealogia', False)
        incluir_estatisticas = request.data.get('incluir_estatisticas', False)
        formato_data = request.data.get('formato_data', 'dd/MM/yyyy')

        # Valida se o usuário tem acesso às propriedades especificadas
        propriedades_usuario = Propriedade.objects.filter(proprietario=request.user)
        
        # Monta a query base com filtros de segurança
        queryset = self.get_queryset()

        if propriedade_id:
            # Verifica se a propriedade especificada pertence ao usuário
            if not propriedades_usuario.filter(id=propriedade_id).exists():
                return Response(
                    {'error': 'Você não tem permissão para exportar dados desta propriedade'},
                    status=status.HTTP_403_FORBIDDEN
                )
            queryset = queryset.filter(propriedade_id=propriedade_id)
        else:
            # Se não especificar propriedade, limita às propriedades do usuário
            queryset = queryset.filter(propriedade__in=propriedades_usuario)

        # Aplica filtros adicionais
        especie_id = request.data.get('especie_id')
        if especie_id:
            queryset = queryset.filter(especie_id=especie_id)

        status_filter = request.data.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        search = request.data.get('search')
        if search:
            queryset = queryset.filter(
                Q(identificacao_unica__icontains=search) |
                Q(nome_registro__icontains=search)
            )

        # Ordena por identificação única
        queryset = queryset.select_related(
            'propriedade', 'especie', 'raca', 'lote_atual', 'pai', 'mae'
        ).order_by('identificacao_unica')

        # Validação de segurança: verifica se todos os animais pertencem ao usuário
        animais_nao_autorizados = queryset.exclude(
            propriedade__in=propriedades_usuario
        ).count()
        
        if animais_nao_autorizados > 0:
            return Response(
                {'error': 'Tentativa de acesso a dados não autorizados detectada'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Cria o workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Animais"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2E8B57")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Define as colunas básicas
        colunas = [
            'ID Único', 'Nome/Registro', 'Propriedade', 'Espécie', 'Raça', 
            'Sexo', 'Data Nascimento', 'Categoria', 'Status', 'Lote Atual',
            'Peso Atual (kg)', 'Idade (meses)'
        ]

        # Adiciona colunas de genealogia se solicitado
        if incluir_genealogia:
            colunas.extend(['Pai', 'Mãe'])

        # Escreve o cabeçalho
        for col_num, column_title in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Escreve os dados dos animais
        row_num = 2
        for animal in queryset:
            # Dados básicos
            dados_animal = [
                animal.identificacao_unica or '',
                animal.nome_registro or '',
                animal.propriedade.nome if animal.propriedade else '',
                animal.especie.nome_display if animal.especie else '',
                animal.raca.nome if animal.raca else '',
                animal.get_sexo_display(),
                animal.data_nascimento.strftime('%d/%m/%Y') if animal.data_nascimento else '',
                animal.get_categoria_display(),
                animal.get_status_display(),
                animal.lote_atual.nome if animal.lote_atual else '',
                animal.get_peso_atual() or '',
                animal.get_idade_meses() or ''
            ]

            # Adiciona dados de genealogia se solicitado
            if incluir_genealogia:
                dados_animal.extend([
                    animal.pai.identificacao_unica if animal.pai else '',
                    animal.mae.identificacao_unica if animal.mae else ''
                ])

            # Escreve a linha
            for col_num, valor in enumerate(dados_animal, 1):
                ws.cell(row=row_num, column=col_num, value=valor)

            row_num += 1

        # Adiciona estatísticas se solicitado
        if incluir_estatisticas:
            # Cria nova aba para estatísticas
            ws_stats = wb.create_sheet("Estatísticas")
            
            # Calcula estatísticas
            total_animais = queryset.count()
            animais_ativos = queryset.filter(status='ativo').count()
            animais_inativos = queryset.filter(status='inativo').count()
            
            # Distribuição por sexo
            machos = queryset.filter(sexo='M').count()
            femeas = queryset.filter(sexo='F').count()
            
            # Distribuição por espécie
            por_especie = queryset.values('especie__nome_display').annotate(
                total=Count('id')
            ).order_by('-total')

            # Escreve estatísticas
            stats_data = [
                ['Estatísticas Gerais', ''],
                ['Total de Animais', total_animais],
                ['Animais Ativos', animais_ativos],
                ['Animais Inativos', animais_inativos],
                ['', ''],
                ['Distribuição por Sexo', ''],
                ['Machos', machos],
                ['Fêmeas', femeas],
                ['', ''],
                ['Distribuição por Espécie', '']
            ]
            
            # Adiciona dados por espécie
            for esp in por_especie:
                stats_data.append([esp['especie__nome_display'], esp['total']])

            # Escreve no worksheet
            for row_num, (label, value) in enumerate(stats_data, 1):
                ws_stats.cell(row=row_num, column=1, value=label)
                if value != '':
                    ws_stats.cell(row=row_num, column=2, value=value)

            # Aplica formatação ao cabeçalho das estatísticas
            for row in [1, 6, 10]:
                cell = ws_stats.cell(row=row, column=1)
                cell.font = header_font
                cell.fill = header_fill

        # Ajusta a largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Gera o arquivo em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepara a resposta
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        propriedade_nome = ''
        if propriedade_id:
            prop = propriedades_usuario.filter(id=propriedade_id).first()
            propriedade_nome = f'_{prop.nome.replace(" ", "_")}' if prop else ''
        
        filename = f'animais_export{propriedade_nome}_{timestamp}.xlsx'
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    @action(detail=False, methods=['post'])
    def importar_planilha(self, request):
        """
        Importa animais de uma planilha Excel (.xlsx) com validações de segurança.
        Só permite importar para propriedades que pertencem ao usuário autenticado.
        """
        import openpyxl
        from django.db import transaction
        from datetime import datetime
        import io

        # Verifica se arquivo foi enviado
        if 'arquivo' not in request.FILES:
            return Response(
                {'error': 'Arquivo não enviado'},
                status=status.HTTP_400_BAD_REQUEST
            )

        arquivo = request.FILES['arquivo']
        
        # Valida extensão do arquivo
        if not arquivo.name.lower().endswith('.xlsx'):
            return Response(
                {'error': 'Formato de arquivo inválido. Use apenas .xlsx'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Valida tamanho do arquivo (máximo 5MB)
        if arquivo.size > 5 * 1024 * 1024:
            return Response(
                {'error': 'Arquivo muito grande. Máximo 5MB permitido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Carrega o arquivo Excel
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            if not wb.worksheets:
                return Response(
                    {'error': 'Planilha vazia ou inválida'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            if len(rows) < 2:
                return Response(
                    {'error': 'Planilha deve conter pelo menos cabeçalho e uma linha de dados'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Primeira linha são os cabeçalhos
            headers = [str(header).strip().lower() if header else '' for header in rows[0]]
            
            # Mapeamento de cabeçalhos em português para nomes técnicos
            # Função auxiliar para normalizar cabeçalhos (remover * e espaços extras)
            def normalizar_header(header):
                return header.replace('*', '').strip().lower()
            
            mapeamento_headers = {
                'id único': 'identificacao_unica',
                'id unico': 'identificacao_unica',
                'identificacao única': 'identificacao_unica',
                'identificacao unica': 'identificacao_unica',
                'identificacao_unica': 'identificacao_unica',
                'identificacao': 'identificacao_unica',
                'nome/registro': 'nome_registro',
                'nome registro': 'nome_registro',
                'nome_registro': 'nome_registro',
                'nome': 'nome_registro',
                'propriedade': 'propriedade',
                'espécie': 'especie',
                'especie': 'especie',
                'raça': 'raca',
                'raca': 'raca',
                'sexo': 'sexo',
                'data nascimento': 'data_nascimento',
                'data_nascimento': 'data_nascimento',
                'data de nascimento': 'data_nascimento',
                'nascimento': 'data_nascimento',
                'categoria': 'categoria',
                'status': 'status',
                'lote atual': 'lote_atual',
                'lote_atual': 'lote_atual',
                'lote': 'lote_atual',
                'peso atual (kg)': 'peso_atual',
                'peso atual': 'peso_atual',
                'peso_atual': 'peso_atual',
                'peso': 'peso_atual',
                'observações': 'observacoes',
                'observacoes': 'observacoes',
                'observação': 'observacoes',
                'obs': 'observacoes',
            }
            
            # Converter cabeçalhos usando o mapeamento com normalização
            headers_normalizados = []
            for header_original in headers:
                # Normalizar header removendo asteriscos e espaços
                header_normalizado = normalizar_header(header_original)
                
                # Buscar no mapeamento usando header normalizado
                header_mapeado = mapeamento_headers.get(header_normalizado, header_original)
                headers_normalizados.append(header_mapeado)
            
            headers = headers_normalizados
            
            # Validar cabeçalhos obrigatórios
            headers_obrigatorios = ['identificacao_unica', 'especie', 'sexo', 'data_nascimento', 'categoria', 'status']
            headers_faltantes = []
            
            for header in headers_obrigatorios:
                if header not in headers:
                    headers_faltantes.append(header)
            
            if headers_faltantes:
                return Response(
                    {'error': f'Colunas obrigatórias ausentes: {", ".join(headers_faltantes)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Validações de segurança - verificar propriedades
            propriedades_usuario = Propriedade.objects.filter(proprietario=request.user)
            if not propriedades_usuario.exists():
                return Response(
                    {'error': 'Usuário não possui propriedades cadastradas'},
                    status=status.HTTP_403_FORBIDDEN
                )

            # Processar linhas de dados
            animais_validos = []
            animais_duplicados = []
            erros = []
            linha_num = 1  # Primeira linha de dados

            with transaction.atomic():
                for row_index, row in enumerate(rows[1:]):  # Pula o cabeçalho
                    linha_num += 1
                    
                    # Pular linhas completamente vazias
                    if not any(row):
                        continue
                    
                    try:
                        # Criar dicionário com dados da linha usando headers mapeados
                        dados_animal = {}
                        for i, valor in enumerate(row):
                            if i < len(headers):  # Usar headers já mapeados
                                header_mapeado = headers[i]
                                valor_processado = str(valor).strip() if valor else ''
                                dados_animal[header_mapeado] = valor_processado
                        
                        # Validações básicas
                        erros_linha = []
                        
                        # Validar identificação única
                        identificacao = dados_animal.get('identificacao_unica', '').strip()
                        if not identificacao:
                            erros_linha.append('Identificação única é obrigatória')
                        elif Animal.objects.filter(identificacao_unica=identificacao).exists():
                            # Animal já existe - pular esta linha sem erro
                            animais_duplicados.append(identificacao)
                            continue

                        # Validar espécie
                        nome_especie = dados_animal.get('especie', '').strip()
                        especie = None
                        if nome_especie:
                            try:
                                # Tentar primeiro pelo nome_display (ex: "Bovino")
                                especie = EspecieAnimal.objects.get(nome_display__iexact=nome_especie)
                            except EspecieAnimal.DoesNotExist:
                                try:
                                    # Tentar pelo nome técnico (ex: "bovino")
                                    especie = EspecieAnimal.objects.get(nome__iexact=nome_especie)
                                except EspecieAnimal.DoesNotExist:
                                    erros_linha.append(f'Espécie "{nome_especie}" não encontrada')
                        else:
                            erros_linha.append('Espécie é obrigatória')

                        # Validar sexo
                        sexo = dados_animal.get('sexo', '').strip().upper()
                        if sexo not in ['M', 'F', 'MACHO', 'FEMEA', 'MASCULINO', 'FEMININO', 'FÊMEA', 'FÉMEA']:
                            erros_linha.append('Sexo deve ser M/F ou Macho/Fêmea')
                        
                        # Normalizar sexo
                        if sexo in ['MACHO', 'MASCULINO']:
                            sexo = 'M'
                        elif sexo in ['FEMEA', 'FEMININO', 'FÊMEA', 'FÉMEA']:
                            sexo = 'F'

                        # Validar data de nascimento
                        data_nascimento_str = dados_animal.get('data_nascimento', '').strip()
                        data_nascimento = None
                        if data_nascimento_str:
                            try:
                                # Tentar diferentes formatos de data
                                formatos = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y']
                                for formato in formatos:
                                    try:
                                        data_nascimento = datetime.strptime(data_nascimento_str, formato).date()
                                        break
                                    except ValueError:
                                        continue
                                
                                if not data_nascimento:
                                    erros_linha.append(f'Formato de data inválido: "{data_nascimento_str}". Use dd/mm/aaaa')
                            except Exception:
                                erros_linha.append(f'Data de nascimento inválida: "{data_nascimento_str}"')
                        else:
                            erros_linha.append('Data de nascimento é obrigatória')

                        # Validar categoria
                        categoria = dados_animal.get('categoria', '').strip().lower()
                        if not categoria:
                            erros_linha.append('Categoria é obrigatória')

                        # Validar status
                        status_animal = dados_animal.get('status', '').strip().lower()
                        if status_animal not in ['ativo', 'inativo', 'vendido', 'morto', 'descartado']:
                            erros_linha.append('Status deve ser "ativo", "inativo", "vendido", "morto" ou "descartado"')

                        # Validar propriedade (se especificada)
                        propriedade_nome = dados_animal.get('propriedade', '').strip()
                        propriedade = None
                        if propriedade_nome:
                            # Buscar propriedade por nome
                            try:
                                propriedade = propriedades_usuario.get(nome__iexact=propriedade_nome)
                            except Propriedade.DoesNotExist:
                                erros_linha.append(f'Propriedade "{propriedade_nome}" não encontrada')
                        else:
                            # Se não especificada, usar a primeira propriedade do usuário
                            propriedade = propriedades_usuario.first()

                        # Validar raça (opcional)
                        raca = None
                        nome_raca = dados_animal.get('raca', '').strip()
                        if nome_raca and especie:
                            try:
                                raca = RacaAnimal.objects.get(nome__iexact=nome_raca, especie=especie)
                            except RacaAnimal.DoesNotExist:
                                # Se não encontrar a raça, apenas avisa mas não impede a importação
                                erros_linha.append(f'Aviso: Raça "{nome_raca}" não encontrada para a espécie "{nome_especie}". Animal será importado sem raça.')

                        # Validar lote (opcional)
                        lote = None
                        lote_nome = dados_animal.get('lote_atual', '').strip()
                        if lote_nome and propriedade:
                            try:
                                lote = Lote.objects.get(nome__iexact=lote_nome, propriedade=propriedade)
                            except Lote.DoesNotExist:
                                # Lote não encontrado é apenas um aviso
                                pass

                        # Validar peso (opcional)
                        peso_atual = None
                        peso_str = dados_animal.get('peso_atual', '').strip()
                        if peso_str:
                            try:
                                peso_atual = float(peso_str.replace(',', '.'))
                                if peso_atual <= 0:
                                    peso_atual = None  # Ignora peso inválido
                            except ValueError:
                                # Peso inválido é ignorado
                                peso_atual = None

                        # Separar erros críticos de avisos
                        erros_criticos = [erro for erro in erros_linha if not erro.startswith('Aviso:')]
                        
                        # Se houver erros críticos nesta linha, adiciona aos erros gerais
                        if erros_criticos:
                            erros.extend([f'Linha {linha_num}: {erro}' for erro in erros_criticos])
                            continue

                        # Criar objeto Animal se todas as validações passaram
                        nome_registro = dados_animal.get('nome_registro', '').strip()
                        if not nome_registro:
                            nome_registro = f"Animal {identificacao}"  # Nome padrão se não informado
                        
                        observacoes = dados_animal.get('observacoes', '').strip()
                        if not observacoes:
                            observacoes = "Importado da planilha Excel"  # Observação padrão
                        
                        animal_data = {
                            'identificacao_unica': identificacao,
                            'nome_registro': nome_registro,
                            'especie': especie,
                            'raca': raca,
                            'sexo': sexo,
                            'data_nascimento': data_nascimento,
                            'categoria': categoria,
                            'status': status_animal,
                            'propriedade': propriedade,
                            'lote_atual': lote,
                            'observacoes': observacoes,
                        }

                        # Criar o animal
                        animal = Animal.objects.create(**animal_data)
                        
                        # Adicionar peso inicial se especificado
                        if peso_atual:
                            # Criar manejo de pesagem
                            manejo_pesagem = Manejo.objects.create(
                                propriedade=propriedade,
                                tipo='pesagem',
                                data_manejo=data_nascimento or timezone.now().date(),
                                observacoes='Peso inicial importado da planilha',
                                usuario=request.user
                            )
                            
                            # Criar a pesagem vinculada ao manejo
                            pesagem = Pesagem.objects.create(
                                animal=animal,
                                manejo=manejo_pesagem,
                                data_pesagem=data_nascimento or timezone.now().date(),
                                peso_kg=peso_atual,
                                observacoes='Peso inicial importado da planilha'
                            )
                            
                            # Associar o animal ao manejo
                            manejo_pesagem.animais.add(animal)

                        animais_validos.append(animal.identificacao_unica)

                    except Exception as e:
                        erros.append(f'Linha {linha_num}: Erro inesperado - {str(e)}')

            # Resultado da importação
            total_processados = linha_num - 1 - len(animais_duplicados)  # Exclui duplicados do total
            
            resultado = {
                'status': 'sucesso' if not erros else ('parcial' if animais_validos else 'erro'),
                'total_registros': linha_num - 1,
                'sucessos': len(animais_validos),
                'erros': len(erros),
                'duplicados': len(animais_duplicados),
                'animais_importados': animais_validos[:10],  # Primeiros 10 para não sobrecarregar
                'animais_duplicados': animais_duplicados[:10],  # Primeiros 10 duplicados
                'mensagens_erro': erros[:20],  # Primeiros 20 erros
            }

            # Mensagens informativas
            mensagens = []
            if animais_validos:
                mensagens.append(f'{len(animais_validos)} animais importados com sucesso')
            
            if animais_duplicados:
                mensagens.append(f'{len(animais_duplicados)} animais ignorados por já existirem')
            
            if erros:
                mensagens.append(f'{len(erros)} erros encontrados durante a importação')
            
            resultado['message'] = ' • '.join(mensagens) if mensagens else 'Nenhum animal foi processado'

            # Determinar status HTTP correto
            if animais_validos:
                # Se há animais importados com sucesso, retorna 201 Created
                http_status = status.HTTP_201_CREATED
            elif animais_duplicados and not erros:
                # Se só há duplicados (sem erros), retorna 200 OK pois o processamento foi bem-sucedido
                http_status = status.HTTP_200_OK
            else:
                # Se há erros ou nenhum processamento, retorna 400 Bad Request
                http_status = status.HTTP_400_BAD_REQUEST

            return Response(resultado, status=http_status)

        except Exception as e:
            return Response(
                {'error': f'Erro ao processar arquivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def template_importacao(self, request):
        """
        Gera template de importação em Excel com headers em português
        """
        import openpyxl
        import openpyxl.utils
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from django.http import HttpResponse
        import io

        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template Importação"

        # Estilos aprimorados
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill("solid", fgColor="2E8B57")  # Verde escuro
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        example_font = Font(size=11)
        example_fill = PatternFill("solid", fgColor="E8F5E8")  # Verde claro
        example_alignment = Alignment(horizontal="center")
        
        required_fill = PatternFill("solid", fgColor="FFE4E1")  # Rosa claro para obrigatórios
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Cabeçalhos em português (conforme mapeamento no código e planilha de exportação)
        headers = [
            'ID Único',              # identificacao_unica
            'Nome/Registro',         # nome_registro  
            'Propriedade',           # propriedade
            'Espécie',               # especie
            'Raça',                  # raca
            'Sexo',                  # sexo
            'Data Nascimento',       # data_nascimento
            'Categoria',             # categoria
            'Status',                # status
            'Lote Atual',           # lote_atual
            'Peso Atual (kg)',      # peso_atual
            'Observações'           # observacoes
        ]
        
        # Headers técnicos correspondentes (linha oculta para referência)
        headers_tecnicos = [
            'identificacao_unica', 'nome_registro', 'propriedade', 'especie', 'raca', 'sexo', 
            'data_nascimento', 'categoria', 'status', 'lote_atual', 'peso_atual', 'observacoes'
        ]

        # Indicadores de campos obrigatórios
        obrigatorios = [True, False, False, True, False, True, True, True, True, False, False, False]

        # Escrever cabeçalhos em português
        for col_num, (header, obrigatorio) in enumerate(zip(headers, obrigatorios), 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = f"{header} {'*' if obrigatorio else ''}"
            cell.font = header_font
            cell.fill = header_fill if obrigatorio else PatternFill("solid", fgColor="4682B4")  # Azul para opcionais
            cell.alignment = header_alignment
            cell.border = border

        # Linha com exemplos práticos
        exemplos = [
            'BOV001',                           # identificacao_unica
            'Vaca da Silva',                    # nome_registro
            'Fazenda Principal',               # propriedade
            'Bovino',                          # especie (aceita: Bovino, Ovino, Caprino, Suíno, Equino)
            'Nelore',                          # raca
            'F',                               # sexo (M/F ou Macho/Fêmea)
            '15/03/2020',                      # data_nascimento
            'vaca',                            # categoria
            'ativo',                           # status (ativo/inativo)
            'Lote 1',                         # lote_atual
            '450.5',                           # peso_atual
            'Animal reprodutor de alta genética'  # observacoes
        ]

        for col_num, (valor, obrigatorio) in enumerate(zip(exemplos, obrigatorios), 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = valor
            cell.font = example_font
            cell.fill = required_fill if obrigatorio else example_fill
            cell.alignment = example_alignment
            cell.border = border

        # Adicionar mais exemplos variados
        exemplos_adicionais = [
            ['OV002', 'Ovelha Branca', 'Fazenda Principal', 'Ovino', 'Santa Inês', 'F', '10/08/2021', 'matriz', 'ativo', 'Lote 2', '65.0', 'Boa produção de leite'],
            ['CAP003', 'Cabra Preta', 'Fazenda Principal', 'Caprino', 'Anglo Nubiano', 'F', '05/12/2021', 'cabra', 'ativo', '', '45.2', 'Animal jovem'],
            ['BOV004', 'Touro Champion', 'Fazenda Principal', 'Bovino', 'Angus', 'M', '20/01/2019', 'reprodutor', 'ativo', 'Lote 1', '850.0', 'Reprodutor premium'],
            ['EQ005', 'Cavalo Veloz', 'Fazenda Principal', 'Equino', 'Mangalarga', 'M', '12/06/2018', 'garanhão', 'ativo', '', '480.0', 'Cavalo de trabalho']
        ]

        for row_idx, exemplo in enumerate(exemplos_adicionais, 3):
            for col_num, (valor, obrigatorio) in enumerate(zip(exemplo, obrigatorios), 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.value = valor
                cell.font = example_font
                cell.fill = PatternFill("solid", fgColor="F0F8FF") if obrigatorio else PatternFill("solid", fgColor="F5F5F5")
                cell.alignment = example_alignment
                cell.border = border

        # Criar aba de instruções melhorada
        ws_instrucoes = wb.create_sheet("📋 Instruções de Uso")
        
        # Estilos para instruções
        title_font = Font(bold=True, color="FFFFFF", size=14)
        title_fill = PatternFill("solid", fgColor="2E8B57")
        subtitle_font = Font(bold=True, size=12, color="2E8B57")
        text_font = Font(size=11)
        required_font = Font(size=11, color="B22222")
        optional_font = Font(size=11, color="4682B4")
        
        instrucoes = [
            ['🚀 GUIA COMPLETO DE IMPORTAÇÃO DE ANIMAIS', '', '', ''],
            ['', '', '', ''],
            ['📋 COMO USAR ESTE TEMPLATE:', '', '', ''],
            ['1. Preencha os dados na aba "Template Importação"', '', '', ''],
            ['2. Use os exemplos como referência', '', '', ''],
            ['3. Campos obrigatórios estão marcados com * e têm fundo rosa', '', '', ''],
            ['4. Campos opcionais têm fundo azul', '', '', ''],
            ['5. Salve o arquivo em formato Excel (.xlsx)', '', '', ''],
            ['6. Importe através do aplicativo AgroNexus', '', '', ''],
            ['', '', '', ''],
            
            ['✅ CAMPOS OBRIGATÓRIOS (devem ser preenchidos):', '', '', ''],
            ['• Identificação Única*', 'ID único do animal no seu sistema', 'Exemplos: BOV001, VACA123, OV045', ''],
            ['• Espécie*', 'Tipo do animal', 'Bovino, Ovino, Caprino, Suíno, Equino', ''],
            ['• Sexo*', 'Sexo do animal', 'M, F, Macho, Fêmea, Masculino, Feminino', ''],
            ['• Data de Nascimento*', 'Data em formato brasileiro', '15/03/2020, 01/12/2021', ''],
            ['• Categoria*', 'Classificação do animal', 'vaca, touro, novilha, bezerro, etc.', ''],
            ['• Status*', 'Situação atual', 'ativo, inativo, vendido, morto', ''],
            ['', '', '', ''],
            
            ['📝 CAMPOS OPCIONAIS (podem ficar vazios):', '', '', ''],
            ['• Nome do Animal', 'Nome ou registro do animal', 'Vaca da Silva, Touro Champion', ''],
            ['• Raça', 'Raça do animal (deve existir no sistema)', 'Nelore, Angus, Santa Inês', ''],
            ['• Peso Atual', 'Peso em quilogramas', '450.5, 65.0, 850.0', ''],
            ['• Propriedade', 'Nome da propriedade (se vazio, usa a primeira)', 'Fazenda Principal, Sítio do João', ''],
            ['• Lote Atual', 'Nome do lote (deve existir na propriedade)', 'Lote 1, Pasto Norte, Área A', ''],
            ['• Observações', 'Informações adicionais sobre o animal', 'Reprodutor, Boa leiteira, Animal jovem', ''],
            ['', '', '', ''],
            
            ['⚠️ REGRAS IMPORTANTES:', '', '', ''],
            ['✓ Identificação única deve ser diferente para cada animal', '', '', ''],
            ['✓ Animais com identificação já existente serão ignorados', '', '', ''],
            ['✓ Use formato de data brasileiro: DD/MM/AAAA', '', '', ''],
            ['✓ Máximo de 1.000 animais por importação', '', '', ''],
            ['✓ Não deixe linhas vazias entre os dados', '', '', ''],
            ['✓ Propriedade e lote devem existir no seu sistema', '', '', ''],
            ['✓ Arquivo deve estar no formato Excel (.xlsx)', '', '', ''],
            ['', '', '', ''],
            
            ['🔍 VALORES ACEITOS POR CAMPO:', '', '', ''],
            ['📌 Espécie:', 'Bovino, Ovino, Caprino, Suíno, Equino', '', ''],
            ['📌 Sexo:', 'M, F, Macho, Fêmea, Masculino, Feminino', '', ''],
            ['📌 Status:', 'ativo, inativo, vendido, morto, descartado', '', ''],
            ['📌 Categoria:', 'Qualquer texto (ex: vaca, touro, bezerro)', '', ''],
            ['📌 Data:', 'DD/MM/AAAA (ex: 15/03/2020, 01/12/2021)', '', ''],
            ['📌 Peso:', 'Números com ou sem decimal (ex: 450, 380.5)', '', ''],
            ['', '', '', ''],
            
            ['❌ ERROS COMUNS E COMO EVITAR:', '', '', ''],
            ['🚫 "Identificação já existe"', '→ Use IDs únicos para cada animal', '', ''],
            ['🚫 "Espécie não encontrada"', '→ Use: Bovino, Ovino, Caprino, Suíno ou Equino', '', ''],
            ['🚫 "Data inválida"', '→ Use formato DD/MM/AAAA (ex: 15/03/2020)', '', ''],
            ['🚫 "Propriedade não encontrada"', '→ Use nome exato ou deixe vazio', '', ''],
            ['🚫 "Arquivo corrompido"', '→ Salve em formato .xlsx (Excel)', '', ''],
            ['', '', '', ''],
            
            ['💡 DICAS PARA SUCESSO:', '', '', ''],
            ['✨ Teste primeiro com poucos animais (5-10)', '', '', ''],
            ['✨ Verifique se raças existem no seu sistema', '', '', ''],
            ['✨ Use identificações simples e organizadas', '', '', ''],
            ['✨ Mantenha backups dos seus dados', '', '', ''],
            ['✨ Importe em lotes pequenos se tiver muitos animais', '', '', ''],
            ['', '', '', ''],
            
            ['📞 SUPORTE:', '', '', ''],
            ['Em caso de dúvidas, contate o suporte técnico', '', '', ''],
            ['ou consulte a documentação do AgroNexus.', '', '', '']
        ]

        # Aplicar estilos e conteúdo
        for row_num, linha in enumerate(instrucoes, 1):
            for col_num, texto in enumerate(linha, 1):
                if texto:  # Só processa células com conteúdo
                    cell = ws_instrucoes.cell(row=row_num, column=col_num)
                    cell.value = texto
                    
                    # Aplicar estilos baseados no conteúdo
                    if texto.startswith('🚀'):
                        cell.font = title_font
                        cell.fill = title_fill
                        cell.alignment = Alignment(horizontal="center")
                    elif texto.startswith(('📋', '✅', '📝', '⚠️', '🔍', '❌', '💡', '📞')):
                        cell.font = subtitle_font
                    elif '*' in texto and col_num == 1:
                        cell.font = required_font
                    elif col_num == 1 and texto.startswith('•'):
                        cell.font = optional_font
                    else:
                        cell.font = text_font

        # Ajustar largura das colunas para melhor visualização
        # Template principal
        larguras_template = [18, 20, 12, 15, 8, 18, 15, 12, 15, 20, 15, 35]
        for col_num, largura in enumerate(larguras_template, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = largura
            
        # Aba de instruções
        ws_instrucoes.column_dimensions['A'].width = 45
        ws_instrucoes.column_dimensions['B'].width = 35
        ws_instrucoes.column_dimensions['C'].width = 30
        ws_instrucoes.column_dimensions['D'].width = 15
        
        # Congelar primeira linha no template
        ws.freeze_panes = 'A2'
        
        # Adicionar validação de dados para alguns campos
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Validação para coluna Sexo
        sexo_validation = DataValidation(
            type="list",
            formula1='"M,F,Macho,Fêmea,Masculino,Feminino"',
            showErrorMessage=True,
            error="Valor inválido! Use: M, F, Macho, Fêmea, Masculino ou Feminino"
        )
        ws.add_data_validation(sexo_validation)
        sexo_validation.add(f'F2:F1000')  # Coluna Sexo
        
        # Validação para coluna Status
        status_validation = DataValidation(
            type="list", 
            formula1='"ativo,inativo,vendido,morto,descartado"',
            showErrorMessage=True,
            error="Valor inválido! Use: ativo, inativo, vendido, morto ou descartado"
        )
        ws.add_data_validation(status_validation)
        status_validation.add(f'I2:I1000')  # Coluna Status
        
        # Validação para coluna Espécie
        especie_validation = DataValidation(
            type="list",
            formula1='"Bovino,Ovino,Caprino,Suíno,Equino"',
            showErrorMessage=True,
            error="Valor inválido! Use: Bovino, Ovino, Caprino, Suíno ou Equino"
        )
        ws.add_data_validation(especie_validation)
        especie_validation.add(f'D2:D1000')  # Coluna Espécie

        # Gerar arquivo em memória
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Resposta HTTP com nome de arquivo mais descritivo
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="AgroNexus_Template_Importacao_Animais.xlsx"'
        
        return response


class LoteViewSet(BaseViewSet):
    """ViewSet para lotes"""
    queryset = Lote.objects.all()
    serializer_class = LoteSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]
    filterset_class = LoteFilter
    search_fields = ['nome', 'descricao', 'criterio_agrupamento']
    ordering_fields = ['nome', 'data_criacao']
    ordering = ['nome']

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'propriedade', 'area_atual'
        )

        # Adiciona campos calculados
        for lote in queryset:
            lote.total_animais = lote.get_total_animais()
            lote.total_ua = lote.get_total_ua()
            lote.peso_medio = lote.get_peso_medio()
            lote.gmd_medio = lote.get_gmd_medio()

        return queryset

    @action(detail=True, methods=['get'])
    def animais(self, request, pk=None):
        """Lista animais do lote"""
        lote = self.get_object()
        animais = lote.animais.filter(status='ativo')
        page = self.paginate_queryset(animais)

        if page is not None:
            serializer = AnimalSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = AnimalSerializer(animais, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def estatisticas(self, request, pk=None):
        """Estatísticas do lote"""
        lote = self.get_object()

        # Estatísticas básicas
        total_animais = lote.get_total_animais()
        total_ua = lote.get_total_ua()
        peso_medio = lote.get_peso_medio()
        gmd_medio = lote.get_gmd_medio()

        # Distribuição por categoria
        por_categoria = lote.animais.filter(status='ativo').values(
            'categoria'
        ).annotate(total=Count('id'))

        # Distribuição por sexo
        por_sexo = lote.animais.filter(status='ativo').values(
            'sexo'
        ).annotate(total=Count('id'))

        return Response({
            'basicas': {
                'total_animais': total_animais,
                'total_ua': total_ua,
                'peso_medio': peso_medio,
                'gmd_medio': gmd_medio
            },
            'distribuicao': {
                'por_categoria': list(por_categoria),
                'por_sexo': list(por_sexo)
            }
        })

    @action(detail=False, methods=['get'])
    def disponivel(self, request):
        """Lista lotes disponíveis para associação"""
        propriedade_id = request.query_params.get('propriedade_id')
        
        queryset = self.get_queryset()
        
        if propriedade_id:
            queryset = queryset.filter(propriedade_id=propriedade_id)
        
        # Retorna informações resumidas dos lotes
        lotes_data = []
        for lote in queryset:
            total_femeas = lote.animais.filter(sexo='F', status='ativo').count()
            lotes_data.append({
                'id': str(lote.id),
                'nome': lote.nome,
                'descricao': lote.descricao,
                'total_animais': lote.get_total_animais(),
                'total_femeas': total_femeas,
                'aptidao': lote.aptidao,
                'finalidade': lote.finalidade,
                'sistema_criacao': lote.sistema_criacao,
            })
        
        return Response(lotes_data)


# ============================================================================
# VIEWSETS DE MANEJO
# ============================================================================

class ManejoViewSet(BaseViewSet):
    """ViewSet para manejos"""
    queryset = Manejo.objects.all()
    serializer_class = ManejoSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]
    filterset_class = ManejoFilter
    search_fields = ['observacoes']
    ordering_fields = ['data_manejo', 'tipo', 'custo_total']
    ordering = ['-data_manejo']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'propriedade', 'lote', 'usuario'
        ).prefetch_related('animais')

    @action(detail=False, methods=['get'])
    def relatorio_custos(self, request):
        """Relatório de custos por período"""
        data_inicio = request.query_params.get('data_inicio')
        data_fim = request.query_params.get('data_fim')

        queryset = self.get_queryset()

        if data_inicio:
            queryset = queryset.filter(data_manejo__gte=data_inicio)
        if data_fim:
            queryset = queryset.filter(data_manejo__lte=data_fim)

        # Agrupa por tipo de manejo
        custos_por_tipo = queryset.values('tipo').annotate(
            total_material=Sum('custo_material'),
            total_pessoal=Sum('custo_pessoal'),
            total_geral=Sum('custo_total'),
            quantidade=Count('id')
        )

        return Response(list(custos_por_tipo))


class PesagemViewSet(BaseViewSet):
    """ViewSet para pesagens"""
    queryset = Pesagem.objects.all()
    serializer_class = PesagemSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_class = PesagemFilter
    search_fields = ['equipamento_usado', 'observacoes']
    ordering_fields = ['data_pesagem', 'peso_kg']
    ordering = ['-data_pesagem']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'animal', 'manejo'
        ).filter(animal__propriedade__proprietario=self.request.user)

    @action(detail=False, methods=['get'])
    def evolucao_rebanho(self, request):
        """Evolução do peso médio do rebanho"""
        queryset = self.get_queryset()

        # Agrupa por mês
        evolucao = queryset.extra(
            select={'mes': "DATE_FORMAT(data_pesagem, '%%Y-%%m')"}
        ).values('mes').annotate(
            peso_medio=Avg('peso_kg'),
            total_pesagens=Count('id')
        ).order_by('mes')

        return Response(list(evolucao))


# ============================================================================
# VIEWSETS DE REPRODUÇÃO
# ============================================================================

class EstacaoMontaViewSet(BaseViewSet):
    """ViewSet para estações de monta"""
    queryset = EstacaoMonta.objects.all()
    serializer_class = EstacaoMontaSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]
    filterset_class = EstacaoMontaFilter
    search_fields = ['nome', 'observacoes']
    ordering_fields = ['nome', 'data_inicio', 'data_fim']
    ordering = ['-data_inicio']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'propriedade'
        ).prefetch_related('lotes_participantes')

    def perform_create(self, serializer):
        """Se propriedade_id não for fornecida, usa a primeira propriedade do usuário"""
        if 'propriedade_id' not in serializer.validated_data or not serializer.validated_data['propriedade_id']:
            # Busca a primeira propriedade do usuário
            primeira_propriedade = self.request.user.propriedades.first()
            if primeira_propriedade:
                serializer.save(propriedade=primeira_propriedade)
            else:
                # Se não tem propriedade, retorna erro
                from rest_framework.exceptions import ValidationError
                raise ValidationError("Usuário não possui propriedades cadastradas.")
        else:
            serializer.save()

    @action(detail=True, methods=['get'])
    def relatorio_reproducao(self, request, pk=None):
        """Relatório de reprodução da estação"""
        estacao = self.get_object()

        # Estatísticas gerais
        total_femeas = estacao.get_total_femeas()
        taxa_prenhez = estacao.get_taxa_prenhez()

        # Inseminações realizadas
        total_inseminacoes = estacao.inseminacao_set.count()

        # Diagnósticos
        diagnosticos = estacao.inseminacao_set.aggregate(
            positivos=Count('diagnosticos', filter=Q(
                diagnosticos__resultado='positivo')),
            negativos=Count('diagnosticos', filter=Q(
                diagnosticos__resultado='negativo')),
            inconclusivos=Count('diagnosticos', filter=Q(
                diagnosticos__resultado='inconclusivo'))
        )

        return Response({
            'estatisticas_gerais': {
                'total_femeas': total_femeas,
                'taxa_prenhez': taxa_prenhez,
                'total_inseminacoes': total_inseminacoes
            },
            'diagnosticos': diagnosticos
        })

    @action(detail=True, methods=['get'])
    def detalhe(self, request, pk=None):
        """Detalhe completo da estação com lotes associados"""
        estacao = self.get_object()
        
        # Lotes associados com informações resumidas
        lotes_data = []
        for lote in estacao.lotes_participantes.all():
            lotes_data.append({
                'id': str(lote.id),
                'nome': lote.nome,
                'descricao': lote.descricao,
                'total_animais': lote.get_total_animais(),
                'total_femeas': lote.animais.filter(sexo='F', status='ativo').count(),
                'aptidao': lote.aptidao,
                'finalidade': lote.finalidade,
            })
        
        return Response({
            'estacao': self.get_serializer(estacao).data,
            'lotes': lotes_data
        })

    @action(detail=True, methods=['post'])
    def associar_lotes(self, request, pk=None):
        """Associar lotes à estação de monta"""
        estacao = self.get_object()
        lote_ids = request.data.get('lote_ids', [])
        
        if not isinstance(lote_ids, list):
            return Response(
                {'error': 'lote_ids deve ser uma lista'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar se todos os lotes pertencem ao usuário
        lotes = Lote.objects.filter(
            id__in=lote_ids,
            propriedade__proprietario=request.user
        )
        
        if len(lotes) != len(lote_ids):
            return Response(
                {'error': 'Alguns lotes não foram encontrados ou não pertencem ao usuário'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Associar lotes à estação
        estacao.lotes_participantes.set(lotes)
        
        return Response({'message': 'Lotes associados com sucesso'})

    @action(detail=True, methods=['get'])
    def dashboard(self, request, pk=None):
        """Dashboard com estatísticas da estação"""
        estacao = self.get_object()
        
        # Contadores básicos
        total_femeas = estacao.get_total_femeas()
        inseminacoes_realizadas = estacao.inseminacao_set.count()
        
        # Diagnósticos
        diagnosticos_realizados = DiagnosticoGestacao.objects.filter(
            inseminacao__estacao_monta=estacao
        ).count()
        
        diagnosticos_positivos = DiagnosticoGestacao.objects.filter(
            inseminacao__estacao_monta=estacao,
            resultado='positivo'
        ).count()
        
        # Partos
        partos_realizados = Parto.objects.filter(
            mae__inseminacoes__estacao_monta=estacao
        ).distinct().count()
        
        # Cálculo de pendências (estimativas baseadas em prazos)
        from datetime import datetime, timedelta
        
        # Inseminações que deveriam ter diagnóstico (30-60 dias)
        data_limite_diagnostico = datetime.now().date() - timedelta(days=30)
        inseminacoes_sem_diagnostico = estacao.inseminacao_set.filter(
            data_inseminacao__lte=data_limite_diagnostico
        ).exclude(
            diagnosticos__isnull=False
        ).count()
        
        # Gestações que deveriam ter parido (período de gestação + margem)
        gestacoes_sem_parto = DiagnosticoGestacao.objects.filter(
            inseminacao__estacao_monta=estacao,
            resultado='positivo'
        ).exclude(
            inseminacao__animal__partos__data_parto__gt=F('data_diagnostico')
        ).count()
        
        # Taxas
        taxa_prenhez = (diagnosticos_positivos / inseminacoes_realizadas * 100) if inseminacoes_realizadas > 0 else 0
        taxa_parto = (partos_realizados / diagnosticos_positivos * 100) if diagnosticos_positivos > 0 else 0
        
        return Response({
            'estacao_id': str(estacao.id),
            'estacao_nome': estacao.nome,
            'total_femeas': total_femeas,
            'inseminacoes_realizadas': inseminacoes_realizadas,
            'diagnosticos_realizados': diagnosticos_realizados,
            'partos_realizados': partos_realizados,
            'inseminacoes_pendentes': 0,  # Placeholder - definir lógica específica
            'diagnosticos_pendentes': inseminacoes_sem_diagnostico,
            'partos_pendentes': gestacoes_sem_parto,
            'taxa_prenhez': round(taxa_prenhez, 2),
            'taxa_parto': round(taxa_parto, 2),
            'progresso_estacao': round((inseminacoes_realizadas / total_femeas * 100) if total_femeas > 0 else 0, 2),
            'evolucao_temporal': []  # Placeholder - implementar dados temporais se necessário
        })


class ProtocoloIATFViewSet(BaseViewSet):
    """ViewSet para protocolos IATF"""
    queryset = ProtocoloIATF.objects.all()
    serializer_class = ProtocoloIATFSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]
    search_fields = ['nome', 'descricao']
    ordering_fields = ['nome', 'duracao_dias']
    ordering = ['nome']


class InseminacaoViewSet(BaseViewSet):
    """ViewSet para inseminações"""
    queryset = Inseminacao.objects.all()
    serializer_class = InseminacaoSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_class = InseminacaoFilter
    search_fields = ['semen_utilizado', 'observacoes']
    ordering_fields = ['data_inseminacao', 'tipo']
    ordering = ['-data_inseminacao']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'animal', 'manejo', 'reprodutor', 'protocolo_iatf', 'estacao_monta'
        ).filter(animal__propriedade__proprietario=self.request.user)

    def perform_create(self, serializer):
        """Cria o manejo relacionado automaticamente"""
        from ...models import Manejo, AnimalManejo
        
        # Validar se o animal pertence ao usuário
        animal_id = serializer.validated_data.get('animal_id')
        try:
            animal = Animal.objects.get(
                id=animal_id,
                propriedade__proprietario=self.request.user
            )
        except Animal.DoesNotExist:
            raise ValidationError("Animal não encontrado ou não pertence ao usuário")
        
        # Criar o manejo primeiro
        manejo = Manejo.objects.create(
            propriedade=animal.propriedade,
            tipo='inseminacao',
            data_manejo=serializer.validated_data['data_inseminacao'],
            observacoes=serializer.validated_data.get('observacoes', ''),
            usuario=self.request.user
        )
        
        # Criar a inseminação com o manejo
        inseminacao = serializer.save(manejo=manejo, animal=animal)
        
        # Relacionar o animal com o manejo
        AnimalManejo.objects.create(
            animal=animal,
            manejo=manejo
        )

    @action(detail=False, methods=['get'])
    def opcoes_cadastro(self, request):
        """Retorna dados necessários para cadastro de inseminação"""
        from datetime import datetime, timedelta
        
        # Data limite para considerar inseminações recentes (últimos 90 dias)
        data_limite = datetime.now().date() - timedelta(days=90)
        
        # Buscar fêmeas que estão "ocupadas" (inseminadas e ainda não disponíveis)
        femeas_ocupadas = set()
        
        # 1. Fêmeas com inseminações recentes sem diagnóstico
        inseminacoes_sem_diagnostico = Inseminacao.objects.filter(
            animal__propriedade__proprietario=request.user,
            data_inseminacao__gte=data_limite,
            diagnosticos__isnull=True
        ).values_list('animal_id', flat=True)
        femeas_ocupadas.update(inseminacoes_sem_diagnostico)
        
        # 2. Fêmeas com diagnóstico positivo que ainda não pariram
        # Buscar diagnósticos positivos
        diagnosticos_positivos = DiagnosticoGestacao.objects.filter(
            inseminacao__animal__propriedade__proprietario=request.user,
            resultado='positivo'
        ).select_related('inseminacao__animal')
        
        for diagnostico in diagnosticos_positivos:
            animal = diagnostico.inseminacao.animal
            # Verificar se houve parto após este diagnóstico
            parto_posterior = Parto.objects.filter(
                mae=animal,
                data_parto__gt=diagnostico.data_diagnostico
            ).exists()
            
            if not parto_posterior:
                femeas_ocupadas.add(animal.id)
        
        # Animais fêmeas disponíveis para inseminação
        femeas = Animal.objects.filter(
            propriedade__proprietario=request.user,
            sexo='F',
            status='ativo'
        ).exclude(
            id__in=femeas_ocupadas
        ).values('id', 'identificacao_unica', 'nome_registro', 'sexo', 'data_nascimento', 'categoria').order_by('identificacao_unica')[:50]

        # Reprodutores machos do usuário (limitado a 20)
        reprodutores = Animal.objects.filter(
            propriedade__proprietario=request.user,
            sexo='M',
            status='ativo'
        ).values('id', 'identificacao_unica', 'nome_registro', 'sexo', 'data_nascimento', 'categoria').order_by('identificacao_unica')[:20]

        # Protocolos IATF
        protocolos = ProtocoloIATF.objects.filter(
            propriedade__proprietario=request.user
        ).values('id', 'nome', 'descricao')

        # Estações de monta ativas
        estacoes = EstacaoMonta.objects.filter(
            propriedade__proprietario=request.user,
            ativa=True
        ).values('id', 'nome', 'data_inicio', 'data_fim')

        return Response({
            'femeas': list(femeas),
            'reprodutores': list(reprodutores),
            'protocolos_iatf': list(protocolos),
            'estacoes_monta': list(estacoes),
            'tipos_inseminacao': [
                {'value': 'natural', 'label': 'Monta Natural'},
                {'value': 'ia', 'label': 'Inseminação Artificial'},
                {'value': 'iatf', 'label': 'IATF'},
            ]
        })

    @action(detail=False, methods=['get'])
    def estatisticas_reproducao(self, request):
        """Estatísticas gerais de reprodução"""
        from datetime import datetime
        
        # Parâmetros de data
        data_inicio = request.query_params.get('data_inicio')
        data_fim = request.query_params.get('data_fim')
        
        # Se não fornecidas, usar o ano atual
        if not data_inicio:
            data_inicio = f"{datetime.now().year}-01-01"
        if not data_fim:
            data_fim = f"{datetime.now().year}-12-31"
        
        # QuerySet base filtrado por período
        queryset = self.get_queryset().filter(
            data_inseminacao__range=[data_inicio, data_fim]
        )
        
        # Contar inseminações
        total_inseminacoes = queryset.count()
        
        # Diagnósticos positivos
        diagnosticos_positivos = DiagnosticoGestacao.objects.filter(
            inseminacao__in=queryset,
            resultado='positivo'
        ).count()
        
        # Partos realizados
        partos_vivos = Parto.objects.filter(
            mae__propriedade__proprietario=self.request.user,
            data_parto__range=[data_inicio, data_fim],
            resultado='nascido_vivo'
        ).count()
        
        # Taxa de prenhez
        taxa_prenhez = 0
        if total_inseminacoes > 0:
            taxa_prenhez = round((diagnosticos_positivos / total_inseminacoes) * 100, 1)
        
        return Response({
            'inseminacoes': total_inseminacoes,
            'diagnosticos_positivos': diagnosticos_positivos,
            'partos_vivos': partos_vivos,
            'taxa_prenhez': taxa_prenhez,
            'periodo': {
                'data_inicio': data_inicio,
                'data_fim': data_fim
            }
        })

    @action(detail=False, methods=['get'])
    def pendentes_diagnostico(self, request):
        """Inseminações pendentes de diagnóstico"""
        from datetime import datetime, timedelta
        
        # Inseminações dos últimos 30-60 dias que não têm diagnóstico
        data_limite_inicio = datetime.now().date() - timedelta(days=60)
        data_limite_fim = datetime.now().date() - timedelta(days=30)
        
        queryset = self.get_queryset().filter(
            data_inseminacao__range=[data_limite_inicio, data_limite_fim]
        ).exclude(
            diagnosticos__isnull=False
        )
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class DiagnosticoGestacaoViewSet(BaseViewSet):
    """ViewSet para diagnósticos de gestação"""
    queryset = DiagnosticoGestacao.objects.all()
    serializer_class = DiagnosticoGestacaoSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_class = DiagnosticoGestacaoFilter
    search_fields = ['metodo', 'observacoes']
    ordering_fields = ['data_diagnostico', 'resultado']
    ordering = ['-data_diagnostico']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'inseminacao', 'manejo'
        ).filter(inseminacao__animal__propriedade__proprietario=self.request.user)

    @action(detail=False, methods=['get'])
    def pendentes_parto(self, request):
        """Gestações pendentes de parto"""
        from datetime import datetime, timedelta
        
        # Diagnósticos positivos com parto previsto nos próximos 30 dias
        data_limite = datetime.now().date() + timedelta(days=30)
        
        queryset = self.get_queryset().filter(
            resultado='positivo',
            data_parto_prevista__lte=data_limite,
            data_parto_prevista__gte=datetime.now().date()
        ).exclude(
            # Excluir os que já tiveram parto registrado
            inseminacao__animal__partos_mae__data_parto__isnull=False
        )
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


class PartoViewSet(BaseViewSet):
    """ViewSet para partos"""
    queryset = Parto.objects.all()
    serializer_class = PartoSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_class = PartoFilter
    search_fields = ['observacoes']
    ordering_fields = ['data_parto', 'resultado', 'peso_nascimento']
    ordering = ['-data_parto']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'mae', 'manejo', 'bezerro'
        ).filter(mae__propriedade__proprietario=self.request.user)


# ============================================================================
# VIEWSETS DE SANIDADE
# ============================================================================

class VacinaViewSet(viewsets.ModelViewSet):
    """ViewSet para vacinas"""
    queryset = Vacina.objects.all()
    serializer_class = VacinaSerializer
    permission_classes = [permissions.IsAuthenticated]
    search_fields = ['nome', 'fabricante', 'doencas_previne']
    ordering_fields = ['nome', 'fabricante']
    ordering = ['nome']

    def get_queryset(self):
        return self.queryset.filter(ativa=True)


class MedicamentoViewSet(viewsets.ModelViewSet):
    """ViewSet para medicamentos"""
    queryset = Medicamento.objects.all()
    serializer_class = MedicamentoSerializer
    permission_classes = [permissions.IsAuthenticated]
    search_fields = ['nome', 'fabricante', 'principio_ativo']
    ordering_fields = ['nome', 'fabricante', 'tipo']
    ordering = ['nome']

    def get_queryset(self):
        return self.queryset.filter(ativo=True)


class VacinacaoViewSet(BaseViewSet):
    """ViewSet para vacinações"""
    queryset = Vacinacao.objects.all()
    serializer_class = VacinacaoSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_class = VacinacaoFilter
    search_fields = ['lote_vacina', 'observacoes']
    ordering_fields = ['manejo__data_manejo', 'data_proxima_dose']
    ordering = ['-manejo__data_manejo']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'manejo', 'vacina'
        ).filter(manejo__propriedade__proprietario=self.request.user)


class AdministracaoMedicamentoViewSet(BaseViewSet):
    """ViewSet para administrações de medicamento"""
    queryset = AdministracaoMedicamento.objects.all()
    serializer_class = AdministracaoMedicamentoSerializer
    permission_classes = [permissions.IsAuthenticated]
    search_fields = ['dosagem_aplicada', 'motivo_aplicacao']
    ordering_fields = ['manejo__data_manejo', 'data_fim_carencia']
    ordering = ['-manejo__data_manejo']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'manejo', 'medicamento'
        ).filter(manejo__propriedade__proprietario=self.request.user)


class CalendarioSanitarioViewSet(BaseViewSet):
    """ViewSet para calendário sanitário"""
    queryset = CalendarioSanitario.objects.all()
    serializer_class = CalendarioSanitarioSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]
    filterset_class = CalendarioSanitarioFilter
    search_fields = ['descricao']
    ordering_fields = ['data_agendada', 'status', 'tipo_manejo']
    ordering = ['data_agendada']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'propriedade', 'vacina', 'medicamento', 'manejo_realizado', 'usuario'
        ).prefetch_related('animais_alvo', 'lotes_alvo')

    @action(detail=False, methods=['get'])
    def proximos_vencimentos(self, request):
        """Retorna próximos vencimentos"""
        dias_antecedencia = int(request.query_params.get('dias', 7))
        data_limite = timezone.now().date() + timedelta(days=dias_antecedencia)

        proximos = self.get_queryset().filter(
            data_agendada__lte=data_limite,
            status='agendado'
        )

        serializer = self.get_serializer(proximos, many=True)
        return Response(serializer.data)


# ============================================================================
# VIEWSETS FINANCEIROS
# ============================================================================

class ContaFinanceiraViewSet(BaseViewSet):
    """ViewSet para contas financeiras"""
    queryset = ContaFinanceira.objects.all()
    serializer_class = ContaFinanceiraSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]
    search_fields = ['nome', 'banco', 'agencia', 'conta']
    ordering_fields = ['nome', 'tipo', 'saldo_inicial']
    ordering = ['nome']

    def get_queryset(self):
        queryset = super().get_queryset().select_related('propriedade')

        # Adiciona saldo atual
        for conta in queryset:
            conta.saldo_atual = conta.get_saldo_atual()

        return queryset


class CategoriaFinanceiraViewSet(BaseViewSet):
    """ViewSet para categorias financeiras"""
    queryset = CategoriaFinanceira.objects.all()
    serializer_class = CategoriaFinanceiraSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]
    search_fields = ['nome', 'descricao']
    ordering_fields = ['nome', 'tipo']
    ordering = ['tipo', 'nome']

    def get_queryset(self):
        return super().get_queryset().filter(ativa=True)


class LancamentoFinanceiroViewSet(BaseViewSet):
    """ViewSet para lançamentos financeiros"""
    queryset = LancamentoFinanceiro.objects.all()
    serializer_class = LancamentoFinanceiroSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]
    filterset_class = LancamentoFinanceiroFilter
    search_fields = ['descricao', 'observacoes']
    ordering_fields = ['data_lancamento', 'valor', 'tipo']
    ordering = ['-data_lancamento']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'propriedade', 'categoria', 'conta_origem', 'conta_destino',
            'manejo_relacionado', 'animal_relacionado', 'usuario'
        )

    @action(detail=False, methods=['get'])
    def fluxo_caixa(self, request):
        """Fluxo de caixa por período"""
        data_inicio = request.query_params.get('data_inicio')
        data_fim = request.query_params.get('data_fim')

        queryset = self.get_queryset()

        if data_inicio:
            queryset = queryset.filter(data_lancamento__gte=data_inicio)
        if data_fim:
            queryset = queryset.filter(data_lancamento__lte=data_fim)

        # Agrupa por mês
        fluxo = queryset.extra(
            select={'mes': "DATE_FORMAT(data_lancamento, '%%Y-%%m')"}
        ).values('mes').annotate(
            entradas=Sum('valor', filter=Q(tipo='entrada')),
            saidas=Sum('valor', filter=Q(tipo='saida')),
            saldo=Sum('valor', filter=Q(tipo='entrada')) -
            Sum('valor', filter=Q(tipo='saida'))
        ).order_by('mes')

        return Response(list(fluxo))


# ============================================================================
# VIEWSETS DE CONFIGURAÇÃO
# ============================================================================

class RelatorioPersonalizadoViewSet(BaseViewSet):
    """ViewSet para relatórios personalizados"""
    queryset = RelatorioPersonalizado.objects.all()
    serializer_class = RelatorioPersonalizadoSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]
    search_fields = ['nome']
    ordering_fields = ['nome', 'tipo', 'data_criacao']
    ordering = ['nome']

    def get_queryset(self):
        return super().get_queryset().select_related(
            'propriedade', 'usuario'
        ).filter(
            Q(usuario=self.request.user) | Q(publico=True)
        )

    def perform_create(self, serializer):
        serializer.save(usuario=self.request.user)


class ConfiguracaoSistemaViewSet(BaseViewSet):
    """ViewSet para configurações do sistema"""
    queryset = ConfiguracaoSistema.objects.all()
    serializer_class = ConfiguracaoSistemaSerializer
    permission_classes = [
        permissions.IsAuthenticated, PropriedadeOwnerPermission]

    def get_queryset(self):
        return super().get_queryset().select_related('propriedade')


# ============================================================================
# VIEWSETS DE HISTÓRICO
# ============================================================================

class HistoricoLoteAnimalViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para histórico de lotes-animal"""
    queryset = HistoricoLoteAnimal.objects.all()
    serializer_class = HistoricoLoteAnimalSerializer
    permission_classes = [permissions.IsAuthenticated]
    search_fields = ['motivo_movimentacao']
    ordering_fields = ['data_entrada', 'data_saida']
    ordering = ['-data_entrada']

    def get_queryset(self):
        return self.queryset.select_related(
            'animal', 'lote', 'usuario'
        ).filter(animal__propriedade__proprietario=self.request.user)


class HistoricoOcupacaoAreaViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para histórico de ocupação de áreas"""
    queryset = HistoricoOcupacaoArea.objects.all()
    serializer_class = HistoricoOcupacaoAreaSerializer
    permission_classes = [permissions.IsAuthenticated]
    search_fields = ['motivo_movimentacao']
    ordering_fields = ['data_entrada', 'data_saida']
    ordering = ['-data_entrada']

    def get_queryset(self):
        return self.queryset.select_related(
            'lote', 'area', 'usuario'
        ).filter(area__propriedade__proprietario=self.request.user)

    def get_queryset(self):
        queryset = super().get_queryset()

        # Adiciona período de ocupação
        for historico in queryset:
            historico.periodo_ocupacao = historico.get_periodo_ocupacao()

        return queryset
